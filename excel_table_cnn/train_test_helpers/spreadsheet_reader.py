import os

import numpy as np
import openpyxl
import torch
from concurrent.futures import ThreadPoolExecutor, as_completed
from tqdm import tqdm

from excel_table_cnn.dl_classification.tensors import parse_table_range
from excel_table_cnn.train_test_helpers.cell_features import feature_order, get_cell_features_xlsx


class SpreadsheetReader:
    def __init__(self, width, height, features_cache_folder):
        self._num_cell_features = 17
        self._map_width = width
        self._map_height = height
        self._features_cache_folder = features_cache_folder
        self._empty_cell = np.array([1.0, 0.0, 0.0, 0.0, 0.0, 1.0, 1.0, 1.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0])


    def processed_file_name(self, file_name: str, spreadsheet_name: str) -> str:
        base_name = os.path.splitext(file_name)[0]
        output_path = os.path.join(self._features_cache_folder, f"{base_name}__{spreadsheet_name}.pt")
        return output_path


    def read_spreadsheet(self, file_path, sheet_name, tables_area):
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name]

        max_row = min(ws.max_row, self._map_height)
        max_col = min(ws.max_column, self._map_width)
        num_feats = self._num_cell_features

        # Pre-allocate flat array
        sheet_tensor = np.zeros((max_row, max_col, num_feats), dtype=np.float32)

        # Avoid small NumPy arrays per cell
        feature_array = np.empty(num_feats, dtype=np.float32)

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col)):
            for col_idx, cell in enumerate(row):
                features = get_cell_features_xlsx(cell)
                for i, key in enumerate(feature_order):
                    feature_array[i] = float(features[key])
                sheet_tensor[row_idx, col_idx] = feature_array

        
        gt_tables = np.array([parse_table_range(area) for area in tables_area], dtype=np.int64)

        # Convert to torch tensors
        sheet_tensor = torch.from_numpy(sheet_tensor)
        gt_tables = torch.from_numpy(gt_tables)

        return sheet_tensor, gt_tables

    def separate_tables_data(self, sheet_tensor, tables_area):
        background_features = sheet_tensor.copy()
        H, W, _ = background_features.shape
        tables_features = []
        for table_area in tables_area:
            min_col, min_row, max_col, max_row = parse_table_range(table_area)
            if min_col >= W or min_row >= H:
                continue
            table_slice = sheet_tensor[min_row:max_row, min_col:max_col].copy()
            h, w, c = table_slice.shape
            if h == 0 or w == 0:
                raise ValueError(f"Table slice has zero height or width: shape {table_slice.shape}")
            tables_features.append(table_slice)
            height, width = table_slice.shape[:2]

            sheet_tensor[min_row:max_row, min_col:max_col] = np.tile(self._empty_cell, (height, width, 1))

        return tables_features, background_features

    def load_dataset_maps(self, labels_df, data_folder):
        """
        Extracts feature maps from a labeled DataFrame using a custom feature extraction function.

        Args:
            labels_df (pd.DataFrame): DataFrame with columns ['sheet_name', 'file_path', 'table_region']
            data_folder (str): Root folder where spreadsheet files are located

        Returns:
            List[Tensor]: A list of feature maps extracted from the dataset
        """
        #tables,  backgrounds = [], []

        for _, row in tqdm(labels_df.iterrows(), total=len(labels_df), desc="Processing sheets"):
            sheet_name = row['sheet_name']
            file_path = os.path.join(data_folder, row['file_path'])
            print(file_path)
            sheet_tensor, gt_tables = self.read_spreadsheet(file_path, sheet_name, row['table_region'])

            save_data = {
                "sheet_tensor": sheet_tensor,
                "gt_tables": gt_tables
            }

            output_path = self.processed_file_name(row['file_path'], sheet_name)
            torch.save(save_data, output_path)

    def parallel_process_maps(self, labels_df, data_folder, max_workers=8):
        """
        Parallel version of load_dataset_maps using ThreadPoolExecutor.

        Args:
            labels_df (pd.DataFrame): DataFrame with columns ['sheet_name', 'file_path', 'table_region']
            data_folder (str): Root folder where spreadsheet files are located
            max_workers (int): Number of threads to use

        Returns:
            None (saves .pt files to disk)
        """

        def process_row(row):
            sheet_name = row['sheet_name']
            file_path = os.path.join(data_folder, row['file_path'])

            sheet_tensor, gt_tables = self.read_spreadsheet(file_path, sheet_name, row['table_region'])
            save_data = {
                "sheet_tensor": sheet_tensor,
                "gt_tables": gt_tables
            }

            output_path = self.processed_file_name(row['file_path'], sheet_name)
            torch.save(save_data, output_path)
            return output_path

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(process_row, row) for _, row in labels_df.iterrows()]

            for _ in tqdm(as_completed(futures), total=len(futures), desc="Processing sheets"):
                pass  # Results are saved to disk


    def pad_feature_map(self, feature_map):
        h, w, d = feature_map.shape

        # Clip dimensions if they exceed max limits
        clipped_h = min(h, self._map_height)
        clipped_w = min(w, self._map_width)

        # Optionally log or warn if clipping occurs
        if h > self._map_height or w > self._map_width:
            print(f"Warning: Feature map clipped from ({h}, {w}) to ({clipped_h}, {clipped_w})")

        # Create padded map filled with _empty_cell values
        resized_map = np.tile(self._empty_cell, (self._map_height, self._map_width, 1))

        # Copy the valid portion of the original map into the top-left corner
        resized_map[:clipped_h, :clipped_w, :] = feature_map[:clipped_h, :clipped_w, :]

        return resized_map

