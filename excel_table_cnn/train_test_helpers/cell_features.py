import os

import pandas as pd
import openpyxl
import torch
from openpyxl.utils import range_boundaries

from excel_table_cnn.dl_classification.tensors import preprocess_features


def get_cell_features_xlsx(cur_cell):
    cell_features = {
        "coordinate": cur_cell.coordinate,
        "is_empty": cur_cell.value is None,
        "is_string": cur_cell.data_type in ["s", "str"],
        "is_merged": type(cur_cell).__name__ == "MergedCell",
        "is_bold": cur_cell.font.b or False,
        "is_italic": cur_cell.font.i or False,
        "left_border": cur_cell.border.left is not None,
        "right_border": cur_cell.border.right is not None,
        "top_border": cur_cell.border.top is not None,
        "bottom_border": cur_cell.border.bottom is not None,
        "is_filled": cur_cell.fill.patternType is not None,
        "horizontal_alignment": cur_cell.alignment.horizontal is not None,
        "left_horizontal_alignment": cur_cell.alignment.horizontal == "left",
        "right_horizontal_alignment": cur_cell.alignment.horizontal == "right",
        "center_horizontal_alignment": cur_cell.alignment.horizontal == "center",
        "wrapped_text": cur_cell.alignment.wrapText or False,
        "indent": cur_cell.alignment.indent != 0,
        "formula": cur_cell.data_type == "f",
    }
    return cell_features


def get_table_features(file_path, sheet_name) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Determine the actual data range
    min_row = 1
    max_row = ws.max_row
    min_col = 1
    max_col = ws.max_column



    data = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            data.append(get_cell_features_xlsx(cell))

    result_df = pd.DataFrame(data)
    return result_df


def process_sheet(self, group, max_rows, max_cols, num_cell_features, non_feature_columns):
        sheet_tensor = torch.zeros((max_rows, max_cols, num_cell_features), dtype=torch.float32)

        # Extract coordinates
        coordinates = np.array([parse_coordinate(coord) for coord in group["coordinate"]])
        row_indices, col_indices = coordinates[:, 0], coordinates[:, 1]

        # Extract and process features
        feature_matrix = np.stack(group.drop(columns=non_feature_columns).apply(preprocess_features, axis=1).to_numpy())

        # Assign values efficiently
        sheet_tensor[row_indices, col_indices, :] = torch.tensor(feature_matrix, dtype=torch.float32)

        return sheet_tensor[:,:,:17]


feature_order = [
    'is_empty',
    'is_string',
    'is_merged',
    'is_bold',
    'is_italic',
    'left_border',
    'right_border',
    'top_border',
    'bottom_border',
    'is_filled',
    'horizontal_alignment',
    'left_horizontal_alignment',
    'right_horizontal_alignment',
    'center_horizontal_alignment',
    'wrapped_text',
    'indent',
    'formula'
]


def get_table_features2(file_path, sheet_name, table_area) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]

    # Determine the actual data range
    # min_row = 1
    # max_row = ws.max_row
    # min_col = 1
    # max_col = ws.max_column

    num_cell_features = 17
    min_col, min_row, max_col, max_row = range_boundaries(table_area)
    sheet_tensor = torch.zeros((max_row - min_row + 1, max_col - min_col + 1, num_cell_features), dtype=torch.float32)

    data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)):
        for col_idx, cell in enumerate(row):
            features = get_cell_features_xlsx(cell)
            feature_matrix = torch.tensor([float(features[key]) for key in feature_order],dtype=torch.float32)
            sheet_tensor[row_idx, col_idx] = feature_matrix

    return sheet_tensor


# def generate_feature_tensor(H, W, device):
#     """
#     Generate a (H, W, 17) tensor where each cell contains the same predefined 17D vector.
#
#     Args:
#         H (int): Height of the output tensor.
#         W (int): Width of the output tensor.
#
#     Returns:
#         torch.Tensor: A tensor of shape (H, W, 17)
#     """
#     # Define the 17D feature vector
#     base_vector = torch.tensor([
#         1.0, 0.0, 0.0, 0.0, 0.0, 1.0, 1.0, 1.0, 1.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
#     ], device=device)
#
#     # Repeat it H * W times and reshape to (H, W, 17)
#     return base_vector.repeat(H * W, 1).view(H, W, 17)

def generate_feature_tensor(H, W, device):
    """
    Efficiently generate a (H, W, 17) tensor with realistic random boolean features,
    including independently randomized border flags.
    """
    N = H * W
    tensor = torch.zeros((N, 17), dtype=torch.float32, device=device)

    # Binary random features (independent)
    tensor[:, 0] = torch.rand(N, device=device) < 0.1   # is_empty
    tensor[:, 1] = 1.0 - tensor[:, 0]                   # is_string = not is_empty
    tensor[:, 2] = torch.rand(N, device=device) < 0.05  # is_merged
    tensor[:, 3] = torch.rand(N, device=device) < 0.2   # is_bold
    tensor[:, 4] = torch.rand(N, device=device) < 0.1   # is_italic

    # Independent random borders
    tensor[:, 5] = torch.rand(N, device=device) < 0.5  # left_border
    tensor[:, 6] = torch.rand(N, device=device) < 0.5  # right_border
    tensor[:, 7] = torch.rand(N, device=device) < 0.5  # top_border
    tensor[:, 8] = torch.rand(N, device=device) < 0.5  # bottom_border

    tensor[:, 9]  = torch.rand(N, device=device) < 0.25  # is_filled
    tensor[:, 14] = torch.rand(N, device=device) < 0.15  # wrapped_text
    tensor[:, 15] = torch.rand(N, device=device) < 0.1   # indent
    tensor[:, 16] = torch.rand(N, device=device) < 0.1   # formula

    # Horizontal alignment: one-hot in indices 10-13
    alignment_indices = torch.multinomial(
        torch.tensor([0.2, 0.3, 0.3, 0.2], device=device),
        num_samples=N,
        replacement=True
    ) + 10  # indices 10, 11, 12, 13
    tensor[torch.arange(N, device=device), alignment_indices] = 1.0

    return tensor.view(H, W, 17)


def extract_feature_maps_from_labels(labels_df, data_folder):
    """
    Extracts feature maps from a labeled DataFrame using a custom feature extraction function.

    Args:
        labels_df (pd.DataFrame): DataFrame with columns ['sheet_name', 'file_path', 'table_region']
        data_folder (str): Root folder where spreadsheet files are located

    Returns:
        List[Tensor]: A list of feature maps extracted from the dataset
    """
    feature_maps = []

    for _, row in labels_df.iterrows():
        sheet_name = row['sheet_name']
        file_path = os.path.join(data_folder, row['file_path'])

        for table_area in row['table_region']:
            features_map = get_table_features2(file_path, sheet_name, table_area)
            feature_maps.append(features_map)

    return feature_maps


